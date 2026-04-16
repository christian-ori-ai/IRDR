from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    return value


def slugify_date(text: str) -> str:
    try:
        return datetime.fromisoformat(text).date().isoformat()
    except ValueError:
        return text


def summary_map(summary_sheet) -> dict[str, Any]:
    data: dict[str, Any] = {}
    for label, value in summary_sheet.iter_rows(min_row=1, max_col=2, values_only=True):
        if label:
            data[str(label)] = cell_value(value)
    return data


def sheet_rows(ws) -> list[dict[str, Any]]:
    headers = [str(cell) if cell is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    rows: list[dict[str, Any]] = []

    for raw_row in ws.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and value != "" for value in raw_row):
            continue

        row = dict(zip(headers, raw_row))
        rows.append(
            {
                "sampleOrder": row.get("Sample Order"),
                "randomDrawOrder": row.get("Random Draw Order"),
                "facility": row.get("Facility") or ws.title,
                "costCenter": row.get("Cost Center"),
                "binCode": row.get("Bin Code"),
                "binDescription": row.get("Bin Description"),
            }
        )

    return rows


def week_payload(workbook_path: Path) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    summary = summary_map(workbook["Summary"])

    week_id = slugify_date(str(summary.get("IRDR weekly sample date", workbook_path.stem)))
    week_label = f"Week of {week_id}"

    facilities: dict[str, Any] = {}
    for facility_name in ("Montgomery", "McKinnon"):
        if facility_name not in workbook.sheetnames:
            continue
        facilities[facility_name] = {
            "population": summary.get(f"{facility_name} population"),
            "sampleSize": summary.get(f"{facility_name} sample"),
            "locations": sheet_rows(workbook[facility_name]),
        }

    return {
        "id": week_id,
        "label": week_label,
        "company": summary.get("Sampling company", "Spokane Stock Yards"),
        "method": summary.get("Method"),
        "confidenceLevel": summary.get("Confidence level"),
        "marginOfError": summary.get("Margin of error target"),
        "excludedCostCenters": [cc.strip() for cc in str(summary.get("Excluded cost centers", "")).split(",") if cc.strip()],
        "sortOrder": summary.get("Sort order"),
        "facilities": facilities,
        "sourceFile": workbook_path.name,
    }


def update_dataset(output_path: Path, week: dict[str, Any]) -> None:
    if output_path.exists():
        existing = json.loads(output_path.read_text(encoding="utf-8"))
    else:
        existing = {"weeks": []}

    weeks = [w for w in existing.get("weeks", []) if w.get("id") != week["id"]]
    weeks.append(week)
    weeks.sort(key=lambda item: item["id"], reverse=True)
    output_path.write_text(json.dumps({"weeks": weeks}, indent=2), encoding="utf-8")


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python scripts/export_irdr_sample.py <workbook.xlsx> [output.json]")
        return 1

    workbook_path = Path(sys.argv[1]).resolve()
    output_path = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else Path(__file__).resolve().parents[1] / "data" / "samples.json"

    week = week_payload(workbook_path)
    update_dataset(output_path, week)
    print(f"Updated {output_path} with week {week['id']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
